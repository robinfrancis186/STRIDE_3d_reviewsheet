from __future__ import annotations

import json
import hashlib
import re
import shutil
import subprocess
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKSPACE = Path("/Users/robinfrancis/Documents/Stride 3d review sheet")
SOURCE_XLSX = Path("/Users/robinfrancis/Downloads/STRIDE_Device_Catalogue_Tracker_v9.xlsx")
OUTPUT_HTML = WORKSPACE / "dashboard" / "stride_device_review_dashboard.html"
ASSET_DIR = WORKSPACE / "dashboard" / "assets"
PRODUCT_ASSET_DIR = ASSET_DIR / "products"
OLD_GENERATED_CONCEPT = ASSET_DIR / "stride_dashboard_concept.png"
COPIED_IMAGE_BY_HASH: dict[str, str] = {}

DIVISIONS = ["Learning", "Elderly Care", "Blind", "Montessori Kits", "Others"]
CATEGORY_LABELS = {
    "Assistive": "Assistive Devices",
    "Cognitive": "Cognitive Devices",
}


def dl(name: str) -> Path:
    return Path("/Users/robinfrancis/Downloads") / name


def resolve_image_path(image_name: str) -> Path:
    path = Path(image_name)
    return path if path.is_absolute() else dl(image_name)


REAL_IMAGES: dict[str, list[str]] = {
    "Minimalist Book Stand / Bookend": ["1f07ebff-b70d-4147-82ff-3a69244754a9.JPG"],
    "Book Stand with Dish": ["4c7ca73e-1f78-4093-b2b4-33fc8ca176cc.JPG"],
    "Digital Huarong Road for Blind": ["02f57089-542a-4331-9f47-17856d99f778.JPG"],
    "Customizable Reaching Tool, Grabber": ["2a59b6e5-d65d-439c-b555-1f70b58e1080.JPG"],
    "Multi Utility Holder / Toothbrush Aid": [
        "5aa92594-d52a-43c4-bba0-942346241cad.JPG",
        "PHOTO-2026-04-28-11-34-05 3.jpg",
    ],
    "Tube Opener": ["5c67bfbe-3f81-4a17-9d6a-fc9fbf9d120d.JPG"],
    "Haptic Memory": ["3ff510d7-c533-4301-8b66-3f9f7bc304e5.JPG", "4a535ed4-c270-4c15-83ea-2a6712071b4e.JPG"],
    "Battleship": ["4aab3d2d-6621-4717-b9ae-ddc75720e2dd.JPG"],
    "Braille Generator": ["2fc07c28-877c-47e7-b565-3d91b1dec5ed.JPG"],
    "Braille Board": ["IMG_20260102_184557.jpg"],
    "Car Door Assist Handle": ["5ffc97ce-fa4c-4535-b108-9a74ae0222be.JPG"],
    "Can Mugger": ["7cb30afa-a855-4814-ae13-f6032f5a2e53.JPG"],
    "Signature Template Aid": ["32c211a1-4e67-427c-a704-a2b150ea5280.JPG"],
    "Pattern Puzzle Block": ["35d5b73b-bfb3-4836-9512-676c43405e5a.JPG", "PHOTO-2026-03-17-15-56-15 2.jpg"],
    "Stacking Toy Collection": [
        "/var/folders/cl/krw8r1cs1wvbqgd9thdtlntc0000gn/T/TemporaryItems/NSIRD_screencaptureui_8XqGKa/Screenshot 2026-07-08 at 11.21.09\u202fam.png",
    ],
    "Learning Slate English Alphabet": ["IMG_9974.PNG"],
    "Learning Slate Malayalam Alphabet": ["IMG_9975.PNG"],
    "4 in a Row Book Storage Edition": ["WhatsApp Image 2026-07-08 at 11.22.46.jpeg"],
    "Bank Heist Brain Teasing Strategy Game": ["WhatsApp Image 2026-07-08 at 11.22.47 (1).jpeg"],
    "Creative Book Style Ludo": ["WhatsApp Image 2026-07-08 at 11.22.47 (2).jpeg"],
    "SEASTRIKE Bookcase Edition": ["WhatsApp Image 2026-07-08 at 11.22.47 (3).jpeg"],
    "Shelf Chess Interactive Book Nook": ["WhatsApp Image 2026-07-08 at 11.22.47.jpeg"],
    "Snakes and Ladders Bookcase Edition": ["WhatsApp Image 2026-07-08 at 11.22.48.jpeg"],
    "Scrabble Book Box": ["WhatsApp Image 2026-07-08 at 11.27.57.jpeg"],
    "Hangman Ultimate Full Game Edition": ["WhatsApp Image 2026-07-08 at 11.28.04.jpeg"],
    "Hangman Gallows Game": ["WhatsApp Image 2026-07-08 at 11.28.05.jpeg"],
    "Mini Sensory Shapes": ["60dc13be-b654-4c7a-9ff8-d257ae1d15c3.JPG"],
    "Tactile Fidget Stones Set of 6": ["images (6).jpeg"],
    "Sensory Textured Floor Mats Circle": ["images (4).jpeg"],
    "Sensory Textured Floor Mats Square": ["images (5).jpeg"],
    "Visually Impaired Calendar Board": ["97fb647e-573d-4f08-a37a-0cb09d0dad00.JPG", "665965ce-0efa-4779-9437-c0144fb7771b.JPG"],
    "Door Knob Handle Adapter": ["171fd781-a468-4d36-b070-eff1e22849b4.JPG"],
    "Playing Cards in Braille": ["230b1ed8-f891-4f0b-9f17-985da282357e.JPG", "63004339-e6ac-4a96-ad0e-3ff8aec04ac6.JPG"],
    "Montessori Lion Puzzle": ["5e92a69d-d889-48ca-8622-c34e24d44978.JPG"],
    "Grocery Bag Carrier Handle": ["1139fe8c-7daf-4d6e-b341-9309cb57d4c9.JPG"],
    "Adjustable Jar Opener": ["2419e8d3-39c8-4100-9652-b5d8cbe4f4e8.JPG"],
    "Sudoku Board & Tiles": ["6768d138-4f32-4413-963b-1bb784e5f94c.JPG"],
    "Jar Opener": ["469581fc-caf8-463c-b238-3d53afa0f9d2.JPG"],
    "Hand Tremor Tool / Brace": ["8272375b-3a27-4039-ab7a-100067e7c5a2.JPG"],
    "Spinning Shapes Children's Toy": ["11670443-41fc-4ef3-8825-6fe64aaa241e.JPG"],
    "Book Page Holder": ["ae464ea2-9c32-4896-9eba-606eee17c222.JPG"],
    "Rolling Hourglass": ["aea24b40-138a-46d6-8427-efe4eba4f155.JPG"],
    "Finger Splint": ["b00b2741-39c2-467e-b295-bc4263f82163.JPG"],
    "Safe Tactile Compass & Ruler": ["c75e111c-46e3-4808-bec2-c0a5013080aa.JPG"],
    "Jar Opener Vacuum Releaser": ["c00792a2-005a-4281-84ed-fdec2ab50645.JPG"],
    "Activity Board": ["cf7ce3d7-bf14-461c-b3db-402122aef69b.JPG"],
    "Pill Box": ["cf7790c1-356a-437c-b95f-2e2af2bf0c10.JPG"],
    "Nalgene Handle": ["dd44550b-0a7d-4ea9-8310-ac3f369ceddb.JPG"],
    "Paper Cup Holder": ["e73359ab-5ea9-438a-a3d5-1eb1655290e9.JPG"],
    "Tangram Blind Low Vision": ["e8289399-9e47-4be8-a062-f8d3c7b0b5ae.JPG"],
    "Maze": ["f53ea5f3-0d0c-40db-8c4c-f8c6a49b1bef.JPG"],
    "Fraction Circle Puzzle": ["PHOTO-2026-03-17-16-52-11.jpg"],
    "Stacking Toy": ["PHOTO-2026-03-17-16-52-11 2.jpg"],
    "Sensory Chew Necklace": ["PHOTO-2026-04-20-16-09-13.jpg"],
    "Oral Motor Toolkit 3-Piece Set": ["images (3).jpeg"],
    "Adaptive T-Bar Oral Motor Chew Tube": ["images (2).jpeg"],
    "Fidget Slug": ["PHOTO-2026-04-20-16-09-23.jpg"],
    "Sensory Rubiks Cube": ["PHOTO-2026-04-21-10-53-17.jpg"],
    "Tetris": ["PHOTO-2026-04-21-10-53-42.jpg"],
    "Tactile Chess Board": ["PHOTO-2026-04-21-11-09-15.jpg"],
    "Adaptive Chew Tube Y Shape": ["images.jpeg"],
    "Adaptive Dual Textured Chew Tube": ["images (1).jpeg"],
    "Micro Travel Chess Set": [
        "stride_dashboard_sources/micro-travel-chess-set-1.webp",
        "stride_dashboard_sources/micro-travel-chess-set-2.webp",
        "stride_dashboard_sources/micro-travel-chess-set-3.webp",
    ],
    "Abacus Model": ["PHOTO-2026-04-28-11-34-03 2.jpg"],
    "Focus-Boosting Eight-Figure Tool": ["PHOTO-2026-04-28-11-34-03 3.jpg"],
    "Low Profile Switch": ["IMG_1913.HEIC"],
    "Learning Scale with Numbers": ["PHOTO-2026-04-28-11-34-03.jpg"],
    "I Know My Shapes Box": ["PHOTO-2026-04-28-11-34-04 2.jpg", "PHOTO-2026-03-17-15-56-15.jpg"],
    "Easy-Open Water Bottle Opener": ["PHOTO-2026-04-28-11-34-04 3.jpg"],
    "Small Pill Popper": ["PHOTO-2026-04-28-11-34-04.jpg"],
    "Bottle Cap Opener": ["PHOTO-2026-04-28-11-34-04 4.jpg"],
    "Thick Handle Cutlery Set": ["PHOTO-2026-04-28-11-34-05 2.jpg"],
    "Can Opener Aid": ["PHOTO-2026-04-28-11-34-06 2.jpg"],
    "Eating Utensil Aid": ["PHOTO-2026-04-28-11-34-06 3.jpg"],
    "Sock Aid": ["PHOTO-2026-04-28-11-34-06.jpg"],
    "Pen Holder Writing Aid": ["PHOTO-2026-04-28-11-34-07.jpg"],
    "Toothbrush Holder": ["IMG_1899.HEIC"],
    "Palm Pen Holder": ["IMG_1905.HEIC"],
    "Shape Sorting Cube": ["PHOTO-2026-03-17-15-56-15.jpg"],
    "Sliding Number Puzzle": ["PHOTO-2026-03-17-15-56-16.jpg", "PHOTO-2026-03-17-15-56-16 2.jpg"],
    "Digit Builders": ["PHOTO-2026-04-28-11-34-02.jpg"],
    "Button Aid": ["PHOTO-2026-04-28-11-34-05.jpg"],
    "Button Aid Product": ["IMG_2268.HEIC"],
}


DEVICE_IMAGE_MAP: dict[str, list[str]] = {
    "Adaptive Chew Tube 'Y' Shape": REAL_IMAGES["Adaptive Chew Tube Y Shape"],
    "Adaptive Dual-Textured Bite Tube Hollow Chew Tool": REAL_IMAGES["Adaptive Dual Textured Chew Tube"],
    "Adaptive Pencil Grip/Holder": REAL_IMAGES["Pen Holder Writing Aid"],
    "Adaptive T-Bar Oral Motor Chew Tube": REAL_IMAGES["Adaptive T-Bar Oral Motor Chew Tube"],
    "Bag Holder": REAL_IMAGES["Grocery Bag Carrier Handle"],
    "Bottle Cap Opener": REAL_IMAGES["Bottle Cap Opener"],
    "Braille": REAL_IMAGES["Braille Board"],
    "Button Aid": REAL_IMAGES["Button Aid Product"],
    "Button Donning Aid for Pants": REAL_IMAGES["Button Aid"],
    "Can Opener Aid - Nail Protector": REAL_IMAGES["Can Opener Aid"],
    "Easy-Open Water Bottle Opener": REAL_IMAGES["Easy-Open Water Bottle Opener"],
    "Eating Utensil Aid": REAL_IMAGES["Eating Utensil Aid"],
    "Fidget Slug": REAL_IMAGES["Fidget Slug"],
    "Low Profile Switch": REAL_IMAGES["Low Profile Switch"],
    "Multi Utility Holder (Toothbrush Aid)": REAL_IMAGES["Multi Utility Holder / Toothbrush Aid"],
    "Oral Motor Toolkit 3-Piece Set": REAL_IMAGES["Oral Motor Toolkit 3-Piece Set"],
    "Palm Pen Holder": REAL_IMAGES["Palm Pen Holder"],
    "Pen Holder Writing Aid": REAL_IMAGES["Pen Holder Writing Aid"],
    "Reading Bar": REAL_IMAGES["Book Page Holder"],
    "Sensory Chew Necklace": REAL_IMAGES["Sensory Chew Necklace"],
    "Sensory Textured Floor Mats - Circle": REAL_IMAGES["Sensory Textured Floor Mats Circle"],
    "Sensory Textured Floor Mats - Square": REAL_IMAGES["Sensory Textured Floor Mats Square"],
    "Small Pill Popper Blister Medicine Tool": REAL_IMAGES["Small Pill Popper"],
    "Sock Aid / Sock Slider Helper": REAL_IMAGES["Sock Aid"],
    "Tactile Fidget Stones - Set of 6": REAL_IMAGES["Tactile Fidget Stones Set of 6"],
    "Thick Handle Cutlery Set (Fork & Knife)": REAL_IMAGES["Thick Handle Cutlery Set"],
    "Toothbrush Holder": REAL_IMAGES["Toothbrush Holder"],
    "Utensil Holder": REAL_IMAGES["Eating Utensil Aid"],
    "Abacus Model": REAL_IMAGES["Abacus Model"],
    "Digit Builders: 7-Segment Display": REAL_IMAGES["Digit Builders"],
    "Focus-Boosting Eight-Figure Tool": REAL_IMAGES["Focus-Boosting Eight-Figure Tool"],
    "Fraction Circle Puzzle": REAL_IMAGES["Fraction Circle Puzzle"],
    "Haptic Memory": REAL_IMAGES["Haptic Memory"],
    "I Know My Shapes Box": REAL_IMAGES["I Know My Shapes Box"],
    "Learning Scale with Numbers": REAL_IMAGES["Learning Scale with Numbers"],
    "Maths Kit 2D": REAL_IMAGES["Pattern Puzzle Block"],
    "Maths Kit 3D": REAL_IMAGES["Fraction Circle Puzzle"],
    "Maths Kit LP": REAL_IMAGES["Abacus Model"],
    "Maths Kit UP": REAL_IMAGES["Learning Scale with Numbers"],
    "Maze": REAL_IMAGES["Maze"],
    "Pattern Puzzle Block": REAL_IMAGES["Pattern Puzzle Block"],
    "Rolling Hourglass": REAL_IMAGES["Rolling Hourglass"],
    "Sensory Rubik's Cube": REAL_IMAGES["Sensory Rubiks Cube"],
    "Shape Sorting Cube for Cognitive Development": REAL_IMAGES["Shape Sorting Cube"],
    "Sliding Number Puzzle - 16": REAL_IMAGES["Sliding Number Puzzle"],
    "Sliding Number Puzzle - 8": REAL_IMAGES["Sliding Number Puzzle"],
    "Stacking Toy - Circle": REAL_IMAGES["Stacking Toy Collection"],
    "Stacking Toy - Cross": REAL_IMAGES["Stacking Toy Collection"],
    "Stacking Toy - Square": REAL_IMAGES["Stacking Toy Collection"],
    "Tactile Chess Board for Visually Impaired Users": REAL_IMAGES["Tactile Chess Board"],
    "Tetris": REAL_IMAGES["Tetris"],
    "Tetris Puzzle Board": REAL_IMAGES["Tetris"],
    "Learning Slate with pencil - Alphabets En": REAL_IMAGES["Learning Slate English Alphabet"],
    "Learning Slate with pencil - Alphabets Malayalam": REAL_IMAGES["Learning Slate Malayalam Alphabet"],
    "Learning Slate with pencil - Shapes": REAL_IMAGES["Shape Sorting Cube"],
    "Learning Slate with pencil - Fruits": REAL_IMAGES["Pattern Puzzle Block"],
    "Learning Slate with pencil - Vegetables": REAL_IMAGES["Pattern Puzzle Block"],
}

DEVICE_IMAGE_OVERRIDES: dict[str, list[str]] = {
    "Adaptive Chew Tube 'Y' Shape": REAL_IMAGES["Adaptive Chew Tube Y Shape"],
    "Adaptive Dual-Textured Bite Tube Hollow Chew Tool": REAL_IMAGES["Adaptive Dual Textured Chew Tube"],
    "Adaptive T-Bar Oral Motor Chew Tube": REAL_IMAGES["Adaptive T-Bar Oral Motor Chew Tube"],
    "Braille": REAL_IMAGES["Braille Board"],
    "Button Aid": REAL_IMAGES["Button Aid Product"],
    "Learning Slate with pencil - Alphabets En": REAL_IMAGES["Learning Slate English Alphabet"],
    "Learning Slate with pencil - Alphabets Malayalam": REAL_IMAGES["Learning Slate Malayalam Alphabet"],
    "Low Profile Switch": REAL_IMAGES["Low Profile Switch"],
    "Oral Motor Toolkit 3-Piece Set": REAL_IMAGES["Oral Motor Toolkit 3-Piece Set"],
    "Palm Pen Holder": REAL_IMAGES["Palm Pen Holder"],
    "Sensory Textured Floor Mats - Circle": REAL_IMAGES["Sensory Textured Floor Mats Circle"],
    "Sensory Textured Floor Mats - Square": REAL_IMAGES["Sensory Textured Floor Mats Square"],
    "Tactile Fidget Stones - Set of 6": REAL_IMAGES["Tactile Fidget Stones Set of 6"],
    "Toothbrush Holder": REAL_IMAGES["Toothbrush Holder"],
    "Stacking Toy - Circle": REAL_IMAGES["Stacking Toy Collection"],
    "Stacking Toy - Cross": REAL_IMAGES["Stacking Toy Collection"],
    "Stacking Toy - Square": REAL_IMAGES["Stacking Toy Collection"],
}

REFERENCE_IMAGE_MAP: dict[str, list[str]] = {
    "Visually Impaired Children's Calendar Board Teaching Aid": REAL_IMAGES["Visually Impaired Calendar Board"],
    "Safe Tactile Compass & Ruler (Braille Version)": REAL_IMAGES["Safe Tactile Compass & Ruler"],
    "Mini Sensory Shapes": REAL_IMAGES["Mini Sensory Shapes"],
    "Playing Cards (Poker) in BRAILLE": REAL_IMAGES["Playing Cards in Braille"],
    "Signature Template aid": REAL_IMAGES["Signature Template Aid"],
    "Beefy Nalgene Handle - Narrow-Mouth Version": REAL_IMAGES["Nalgene Handle"],
    "Door knob to handle adapter": REAL_IMAGES["Door Knob Handle Adapter"],
    "Tangram set for the blind and low vision": REAL_IMAGES["Tangram Blind Low Vision"],
    "Sudoku Board & Tiles": REAL_IMAGES["Sudoku Board & Tiles"],
    "Dual-View (Large Print/Braille) Playing Cards": REAL_IMAGES["Playing Cards in Braille"],
    "Tube Opener": REAL_IMAGES["Tube Opener"],
    "Digital Huarong Road for Blind": REAL_IMAGES["Digital Huarong Road for Blind"],
    "Haptic Memory - Tactile Memory Game": REAL_IMAGES["Haptic Memory"],
    "Paper cups holder for hand support": REAL_IMAGES["Paper Cup Holder"],
    "Braille generator - Any text": REAL_IMAGES["Braille Generator"],
    "Customizable Reaching Tool, Grabber": REAL_IMAGES["Customizable Reaching Tool, Grabber"],
    "Hand tremor tool / brace": REAL_IMAGES["Hand Tremor Tool / Brace"],
    "Can Mugger - Insulated Drinking in Style": REAL_IMAGES["Can Mugger"],
    "Jar Opener": REAL_IMAGES["Jar Opener"],
    "Car Door Assist Handle": REAL_IMAGES["Car Door Assist Handle"],
    "Adjustable jar opener": REAL_IMAGES["Adjustable Jar Opener"],
    "Grocery Bag Carrier Handle": REAL_IMAGES["Grocery Bag Carrier Handle"],
    "Activity board for babies": REAL_IMAGES["Activity Board"],
    "Jar Opener/Vacuum Releaser": REAL_IMAGES["Jar Opener Vacuum Releaser"],
    "Spinning Shapes Children's Toy | Fine Motor Skills": REAL_IMAGES["Spinning Shapes Children's Toy"],
    "FINGER SPLINT - thermoformable little finger splint": REAL_IMAGES["Finger Splint"],
    "Battleship - Naval Strategy Game": REAL_IMAGES["Battleship"],
}

SUPPLEMENTAL_REFERENCES: list[dict[str, Any]] = [
    {
        "siNo": "S1",
        "name": "Minimalist Book Stand / Bookend",
        "designer": "fifindr",
        "platformCategory": "Tools",
        "suggestedSetting": "Special education and assistive learning",
        "description": "Compact bookend and reading support reference captured from the supplied real screenshot set.",
        "downloads": "1.9K",
        "likes": 485,
        "datePublished": "2024-11-30",
        "images": REAL_IMAGES["Minimalist Book Stand / Bookend"],
    },
    {
        "siNo": "S2",
        "name": "Book Stand with Dish",
        "designer": "Neruson",
        "platformCategory": "Tools",
        "suggestedSetting": "Special education and assistive learning",
        "description": "Book stand and dish organizer reference retained from the supplied real image set.",
        "downloads": 50,
        "likes": 21,
        "datePublished": "2025-07-07",
        "images": REAL_IMAGES["Book Stand with Dish"],
    },
    {
        "siNo": "S3",
        "name": "Lion 3D Puzzle | Montessori Educational Game",
        "designer": "mcgiver87",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Anganwadis",
        "description": "Montessori-style animal puzzle reference for fine motor skills, matching, and early learning review.",
        "downloads": 422,
        "likes": 510,
        "datePublished": "",
        "images": REAL_IMAGES["Montessori Lion Puzzle"],
    },
    {
        "siNo": "S4",
        "name": "Pill box weekly - two variants",
        "designer": "motorola2001",
        "platformCategory": "Household",
        "suggestedSetting": "Elderly care and daily living",
        "description": "Weekly pill organization reference for medication routine support and elderly-care review.",
        "downloads": "2.7K",
        "likes": 753,
        "datePublished": "2024-02-26",
        "images": REAL_IMAGES["Pill Box"],
    },
    {
        "siNo": "S5",
        "name": "Micro travel chess set",
        "designer": "Ashwood",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Special education and assistive learning",
        "description": "Compact playable chess set reference measuring around 82mm by 79mm with a low-profile travel format.",
        "downloads": "6.5K",
        "likes": "3.6K",
        "datePublished": "",
        "images": REAL_IMAGES["Micro Travel Chess Set"],
    },
    {
        "siNo": "S6",
        "name": "4 in a Row – Book Storage Edition",
        "designer": "Huba",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Compact Connect Four style game redesigned with hidden book-style storage for classroom shelves and home learning corners.",
        "downloads": 287,
        "likes": 278,
        "datePublished": "2026-01-07",
        "images": REAL_IMAGES["4 in a Row Book Storage Edition"],
    },
    {
        "siNo": "S7",
        "name": "Bank Heist - Brain Teasing Strategy Game",
        "designer": "ozarkexpeditions",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Two-player strategy puzzle game with rotating dials and number logic, suitable for problem-solving review and cognitive challenge screening.",
        "downloads": "3.7K",
        "likes": "3.9K",
        "datePublished": "",
        "images": REAL_IMAGES["Bank Heist Brain Teasing Strategy Game"],
    },
    {
        "siNo": "S8",
        "name": "Creative Book-Style Ludo",
        "designer": "lovelysister",
        "platformCategory": "Hobby & DIY",
        "suggestedSetting": "Anganwadis",
        "description": "Fold-out book-style Ludo board that combines shelf storage with a simple social board game for group play.",
        "downloads": 46,
        "likes": 69,
        "datePublished": "",
        "images": REAL_IMAGES["Creative Book Style Ludo"],
    },
    {
        "siNo": "S9",
        "name": "SEASTRIKE - Bookcase Edition",
        "designer": "Ritz",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Bookcase-format naval combat game with standing boards and reusable tiles, useful as a cognitive strategy and sequencing reference.",
        "downloads": 513,
        "likes": 499,
        "datePublished": "",
        "images": REAL_IMAGES["SEASTRIKE Bookcase Edition"],
    },
    {
        "siNo": "S10",
        "name": "Shelf Chess - interactive book nook",
        "designer": "Schrøder",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Pass-and-play chessboard designed to live on a bookshelf, using snap-in pieces and a compact book-nook format.",
        "downloads": 29,
        "likes": 56,
        "datePublished": "",
        "images": REAL_IMAGES["Shelf Chess Interactive Book Nook"],
    },
    {
        "siNo": "S11",
        "name": "SNAKES & LADDERS - Bookcase Edition",
        "designer": "Ritz",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Anganwadis",
        "description": "Bookcase-edition Snakes and Ladders board with large pieces and a familiar early-learning game format for group engagement.",
        "downloads": 304,
        "likes": 293,
        "datePublished": "",
        "images": REAL_IMAGES["Snakes and Ladders Bookcase Edition"],
    },
    {
        "siNo": "S12",
        "name": "Scrabble Book Box (Version 1)",
        "designer": "BennyCrazyFist",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Storage box concept for a Scrabble board and tiles, relevant as a literacy-game organization and transport reference.",
        "downloads": 26,
        "likes": 25,
        "datePublished": "2026-04-13",
        "images": REAL_IMAGES["Scrabble Book Box"],
    },
    {
        "siNo": "S13",
        "name": "Hangman Ultimate Full Game Edition",
        "designer": "LumeMe",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Full physical Hangman game with alphabet pegs and scoring layout, useful as a spelling and word-recall reference.",
        "downloads": "2K",
        "likes": "3.7K",
        "datePublished": "",
        "images": REAL_IMAGES["Hangman Ultimate Full Game Edition"],
    },
    {
        "siNo": "S14",
        "name": "Hangman game / Gallows game",
        "designer": "Mathias",
        "platformCategory": "Toys & Games",
        "suggestedSetting": "Schools",
        "description": "Simplified hanging figure game model with printed joints, relevant as a compact word-game and tabletop activity reference.",
        "downloads": "3.8K",
        "likes": "2.4K",
        "datePublished": "",
        "images": REAL_IMAGES["Hangman Gallows Game"],
    },
]


def slugify(value: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return text[:72] or "asset"


def is_checked(value: Any) -> bool:
    return str(value or "").strip() in {"\u2714", "Y", "Yes", "TRUE", "True", "1"}


def parse_metric(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip().upper().replace(",", "")
    if not text:
        return None
    multiplier = 1000 if text.endswith("K") else 1
    if text.endswith("K"):
        text = text[:-1]
    try:
        return int(float(text) * multiplier)
    except ValueError:
        return None


def infer_division(name: str, target: str | None, description: str | None = None) -> str:
    n = (name or "").lower()
    t = (target or "").lower()
    d = (description or "").lower()
    combined = f"{n} {t} {d}"
    if not target:
        return "Others"
    if any(term in combined for term in ["braille", "blind", "low vision", "visually impaired", "tactile chess", "haptic memory", "reading bar"]):
        return "Blind"
    if "elderly" in t or any(term in n for term in ["opener", "button", "sock", "pill", "utensil", "cutlery", "toothbrush", "bag holder"]):
        return "Elderly Care"
    if "anganwadi" in t or any(term in n for term in ["stacking toy", "shape sorting", "shapes box", "textured floor", "pattern puzzle", "maze", "tetris"]):
        return "Montessori Kits"
    if any(term in t for term in ["education", "learning", "schools", "primary"]) or any(term in n for term in ["abacus", "digit", "fraction", "maths", "number", "scale", "pencil", "pen", "puzzle", "chew", "fidget"]):
        return "Learning"
    return "Others"


def readiness_status(prod: bool, proto: bool, design: bool) -> str:
    if prod:
        return "Production Ready"
    if proto:
        return "Prototype Ready"
    if design:
        return "Design Ready"
    return "Needs Design"


def next_stage(readiness: str) -> str:
    if readiness == "Design Ready":
        return "Prototype Review"
    if readiness == "Prototype Ready":
        return "Production Review"
    if readiness == "Production Ready":
        return "Field Deployment"
    return "Design Completion"


def review_focus(division: str, status: str, name: str) -> str:
    n = name.lower()
    if division == "Blind":
        return "Validate tactile readability, orientation cues, low-vision contrast, and safe edge finishing."
    if division == "Elderly Care":
        return "Review grip comfort, required hand strength, hygiene, durability, and safe daily handling."
    if division == "Montessori Kits":
        return "Check child-safe edges, part size, color coding, classroom durability, and supervised-use notes."
    if any(term in n for term in ["chew", "oral", "bite"]):
        return "Confirm material safety, cleaning process, bite-risk controls, and supervised-use guidance."
    if status == "Production Ready":
        return "Ready for reviewer sign-off, documentation, and field-use image capture."
    return "Confirm use-case fit, print settings, accessibility value, and field-trial priority."


def copy_real_images(image_names: list[str], asset_prefix: str) -> list[dict[str, str]]:
    copied: list[dict[str, str]] = []
    PRODUCT_ASSET_DIR.mkdir(parents=True, exist_ok=True)
    for image_name in image_names:
        source = resolve_image_path(image_name)
        if not source.exists():
            continue
        source_ext = source.suffix.lower()
        ext = source_ext if source_ext in {".jpg", ".jpeg", ".png", ".webp"} else ".jpg"
        source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
        dest_name = COPIED_IMAGE_BY_HASH.get(source_hash)
        if dest_name is None:
            dest_name = f"{slugify(source.stem)}-{source_hash[:8]}{ext}"
            dest = PRODUCT_ASSET_DIR / dest_name
            if source_ext in {".jpg", ".jpeg", ".png", ".webp"}:
                shutil.copy2(source, dest)
            else:
                subprocess.run(
                    ["sips", "-s", "format", "jpeg", str(source), "--out", str(dest)],
                    check=True,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
            COPIED_IMAGE_BY_HASH[source_hash] = dest_name
        copied.append(
            {
                "src": f"assets/products/{dest_name}",
                "sourceFile": str(source),
                "fileName": source.name,
            }
        )
    return copied


def image_assignment(name: str, fallback_key: str, mapping: dict[str, list[str]], division: str) -> tuple[list[dict[str, str]], str]:
    cleaned = name.strip()
    image_names = mapping.get(cleaned)
    match_quality = "Supplied real image"
    if image_names is None:
        image_names = mapping.get(cleaned.rstrip())
    if image_names is None:
        image_names = {
            "Blind": REAL_IMAGES["Braille Generator"],
            "Elderly Care": REAL_IMAGES["Grocery Bag Carrier Handle"],
            "Montessori Kits": REAL_IMAGES["Pattern Puzzle Block"],
            "Learning": REAL_IMAGES["Abacus Model"],
            "Others": REAL_IMAGES["I Know My Shapes Box"],
        }[division]
        match_quality = "Related supplied real image"
    images = copy_real_images(image_names, fallback_key)
    if not images and match_quality == "Supplied real image":
        fallback_images = {
            "Blind": REAL_IMAGES["Braille Generator"],
            "Elderly Care": REAL_IMAGES["Grocery Bag Carrier Handle"],
            "Montessori Kits": REAL_IMAGES["Pattern Puzzle Block"],
            "Learning": REAL_IMAGES["Abacus Model"],
            "Others": REAL_IMAGES["I Know My Shapes Box"],
        }[division]
        images = copy_real_images(fallback_images, fallback_key)
        match_quality = "Related supplied real image"
    return images, match_quality if images else "Image missing from supplied files"


def load_existing_payload() -> dict[str, Any]:
    if not OUTPUT_HTML.exists():
        raise FileNotFoundError(
            f"{SOURCE_XLSX} is unavailable and no existing dashboard payload was found at {OUTPUT_HTML}"
        )
    html = OUTPUT_HTML.read_text(encoding="utf-8")
    match = re.search(r"const payload = (.*?);\n\s*const ", html, flags=re.S)
    if not match:
        raise ValueError(f"Could not locate embedded dashboard payload in {OUTPUT_HTML}")
    return json.loads(match.group(1))


def append_supplemental_references(references: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_reference_names = {row["name"].casefold() for row in references}
    for supplemental in SUPPLEMENTAL_REFERENCES:
        if supplemental["name"].casefold() in existing_reference_names:
            continue
        division = infer_division(supplemental["name"], supplemental["suggestedSetting"], supplemental["description"])
        images = copy_real_images(supplemental["images"], f"supplemental-{supplemental['siNo']}-{supplemental['name']}")
        references.append(
            {
                "id": f"s-{supplemental['siNo']}",
                "siNo": supplemental["siNo"],
                "name": supplemental["name"],
                "designer": supplemental["designer"],
                "platformCategory": supplemental["platformCategory"],
                "suggestedSetting": supplemental["suggestedSetting"],
                "description": supplemental["description"],
                "downloads": supplemental["downloads"],
                "likes": supplemental["likes"],
                "downloadsNumeric": parse_metric(supplemental["downloads"]),
                "likesNumeric": parse_metric(supplemental["likes"]),
                "datePublished": supplemental["datePublished"],
                "division": division,
                "images": images,
                "imageSource": "Supplemental supplied real image",
            }
        )
        existing_reference_names.add(supplemental["name"].casefold())
    return references


def apply_device_image_overrides(devices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    for device in devices:
        image_names = DEVICE_IMAGE_OVERRIDES.get(device.get("name", ""))
        if image_names is None:
            continue
        images = copy_real_images(image_names, f"device-{device.get('siNo', '')}-{device.get('name', '')}")
        if not images:
            continue
        device["images"] = images
        device["imageSource"] = "Supplied real image"
        device["qaFlags"] = [flag for flag in device.get("qaFlags", []) if flag != "Image missing from supplied files"]
    return devices


def load_catalogue_from_existing_payload() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    payload = load_existing_payload()
    devices = apply_device_image_overrides(list(payload.get("devices", [])))
    references = append_supplemental_references(list(payload.get("references", [])))
    old_metrics = payload.get("metrics", {})
    old_qa = old_metrics.get("qa", {})
    review_items = [*devices, *[reference_to_review_item(row) for row in references]]
    metrics = summarize(
        review_items,
        devices,
        references,
        old_qa.get("sourceDashboardTotal", old_metrics.get("deviceCount", len(devices))),
        old_qa.get("duplicateRowsRemoved", 0),
    )
    metrics["sourceWorkbook"] = f"{SOURCE_XLSX} (workbook unavailable; rebuilt from existing dashboard payload)"
    return devices, references, review_items, metrics


def load_catalogue() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    if not SOURCE_XLSX.exists():
        return load_catalogue_from_existing_payload()

    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Device Catalogue"]
    dashboard_ws = wb["Dashboard"]
    reference_ws = wb["MakerWorld Reference"]

    raw_rows: list[dict[str, Any]] = []
    for row_idx in range(5, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 8)]
        if not any(value is not None for value in values):
            continue
        si_no, name, category, prod, proto, design, target = values
        raw_rows.append(
            {
                "row": row_idx,
                "siNo": si_no,
                "name": str(name or "").strip(),
                "category": str(category or "").strip(),
                "productionReady": is_checked(prod),
                "prototypeReady": is_checked(proto),
                "designReady": is_checked(design),
                "targetSetting": str(target).strip() if target else "",
            }
        )

    seen_raw_rows: set[tuple[Any, ...]] = set()
    deduped_rows: list[dict[str, Any]] = []
    duplicate_rows_removed = 0
    for row in raw_rows:
        row_key = (
            row["siNo"],
            row["name"].casefold(),
            row["category"].casefold(),
            row["productionReady"],
            row["prototypeReady"],
            row["designReady"],
            row["targetSetting"].casefold(),
        )
        if row_key in seen_raw_rows:
            duplicate_rows_removed += 1
            continue
        seen_raw_rows.add(row_key)
        deduped_rows.append(row)
    raw_rows = deduped_rows

    id_counts = Counter(row["siNo"] for row in raw_rows)
    id_occurrences: Counter[Any] = Counter()
    devices: list[dict[str, Any]] = []
    for row in raw_rows:
        source_si_no = row["siNo"]
        id_occurrences[source_si_no] += 1
        if id_counts[source_si_no] > 1:
            suffix = chr(64 + id_occurrences[source_si_no]) if id_occurrences[source_si_no] <= 26 else f"-{id_occurrences[source_si_no]}"
            display_si_no = f"{source_si_no}{suffix}"
        else:
            display_si_no = str(source_si_no)
        category = row["category"]
        division = infer_division(row["name"], row["targetSetting"])
        status = readiness_status(row["productionReady"], row["prototypeReady"], row["designReady"])
        images, match_quality = image_assignment(row["name"], f"device-{display_si_no}-{row['name']}", DEVICE_IMAGE_MAP, division)
        qa_flags: list[str] = []
        if not row["targetSetting"]:
            qa_flags.append("Missing target setting")
        if match_quality != "Supplied real image":
            qa_flags.append(match_quality)
        devices.append(
            {
                **row,
                "sourceSiNo": source_si_no,
                "siNo": display_si_no,
                "id": f"d-{row['row']}",
                "categoryLabel": CATEGORY_LABELS.get(category, category),
                "division": division,
                "readiness": status,
                "nextStage": next_stage(status),
                "readinessRank": {"Production Ready": 3, "Prototype Ready": 2, "Design Ready": 1}.get(status, 0),
                "images": images,
                "imageSource": match_quality,
                "reviewFocus": review_focus(division, status, row["name"]),
                "qaFlags": qa_flags,
                "about": about_device(row["name"], division, row["targetSetting"], status),
            }
        )
    devices = apply_device_image_overrides(devices)

    references: list[dict[str, Any]] = []
    for row_idx in range(5, reference_ws.max_row + 1):
        values = [reference_ws.cell(row_idx, col_idx).value for col_idx in range(1, 10)]
        if not any(value is not None for value in values):
            continue
        si_no, name, designer, platform, target, description, downloads, likes, published = values
        name_text = str(name or "").strip()
        division = infer_division(name_text, str(target or ""), str(description or ""))
        images, match_quality = image_assignment(name_text, f"reference-{si_no}-{name_text}", REFERENCE_IMAGE_MAP, division)
        references.append(
            {
                "id": f"r-{row_idx}",
                "siNo": si_no,
                "name": name_text,
                "designer": str(designer or "").strip(),
                "platformCategory": str(platform or "").strip(),
                "suggestedSetting": str(target or "").strip(),
                "description": str(description or "").strip(),
                "downloads": downloads,
                "likes": likes,
                "downloadsNumeric": parse_metric(downloads),
                "likesNumeric": parse_metric(likes),
                "datePublished": str(published or "").strip(),
                "division": division,
                "images": images,
                "imageSource": match_quality,
            }
        )

    references = append_supplemental_references(references)

    review_items = [*devices, *[reference_to_review_item(row) for row in references]]
    metrics = summarize(review_items, devices, references, dashboard_ws["B6"].value, duplicate_rows_removed)
    return devices, references, review_items, metrics


def about_device(name: str, division: str, target: str, status: str) -> str:
    setting = target or "an unassigned target setting"
    return (
        f"{name} is grouped under {division} for {setting}. "
        f"It is currently marked {status.lower()} in the Excel catalogue. "
        f"The review should confirm user need, print feasibility, safety, and whether the item should advance to {next_stage(status)}."
    )


def category_for_review_item(name: str, division: str, platform: str = "", description: str = "") -> str:
    combined = f"{name} {division} {platform} {description}".lower()
    if division in {"Blind", "Elderly Care"}:
        return "Assistive Devices"
    if any(term in combined for term in ["braille", "blind", "low vision", "assist", "aid", "opener", "handle", "holder", "splint", "tremor"]):
        return "Assistive Devices"
    return "Cognitive Devices"


def about_reference(row: dict[str, Any]) -> str:
    target = row["suggestedSetting"] or "review assignment"
    description = row["description"] or "Supplied real-image reference for catalogue review."
    return (
        f"{row['name']} is an additional real-image reference grouped under {row['division']} for {target}. "
        f"{description} The review should decide whether it should become a STRIDE device candidate, remain pending, or be disposed."
    )


def reference_to_review_item(row: dict[str, Any]) -> dict[str, Any]:
    category_label = category_for_review_item(row["name"], row["division"], row["platformCategory"], row["description"])
    qa_flags: list[str] = []
    if not row["images"]:
        qa_flags.append("Image missing from supplied files")
    if row["imageSource"] == "Related supplied real image":
        qa_flags.append(row["imageSource"])
    return {
        "id": f"ref-{row['id']}",
        "source": "Reference Image",
        "sourceSiNo": row["siNo"],
        "siNo": str(row["siNo"]) if str(row["siNo"]).startswith("S") else f"R{row['siNo']}",
        "name": row["name"],
        "category": "Assistive" if category_label == "Assistive Devices" else "Cognitive",
        "categoryLabel": category_label,
        "division": row["division"],
        "readiness": "Reference Review",
        "nextStage": "Candidate Screening",
        "readinessRank": 0,
        "targetSetting": row["suggestedSetting"],
        "designer": row["designer"],
        "platformCategory": row["platformCategory"],
        "description": row["description"],
        "downloads": row["downloads"],
        "likes": row["likes"],
        "datePublished": row["datePublished"],
        "images": row["images"],
        "imageSource": row["imageSource"],
        "reviewFocus": review_focus(row["division"], "Needs Design", row["name"]),
        "qaFlags": qa_flags,
        "about": about_reference(row),
    }


def summarize(
    review_items: list[dict[str, Any]],
    devices: list[dict[str, Any]],
    references: list[dict[str, Any]],
    source_dashboard_total: Any,
    duplicate_rows_removed: int,
) -> dict[str, Any]:
    category_counts = Counter(item["categoryLabel"] for item in review_items)
    division_counts = Counter(item["division"] for item in review_items)
    readiness_counts = Counter(item["readiness"] for item in review_items)
    by_category_division: dict[str, dict[str, int]] = defaultdict(lambda: {division: 0 for division in DIVISIONS})
    for item in review_items:
        by_category_division[item["categoryLabel"]][item["division"]] += 1

    assigned_images = sum(1 for item in review_items if item["imageSource"] in {"Supplied real image", "Supplemental supplied real image"})
    with_images = sum(1 for item in review_items if item["images"])
    reference_with_images = sum(1 for ref in references if ref["images"])
    source_duplicate_ids = sorted(
        {device["sourceSiNo"] for device in devices if sum(1 for item in devices if item["sourceSiNo"] == device["sourceSiNo"]) > 1}
    )
    display_duplicate_ids = sorted(
        {item["siNo"] for item in review_items if sum(1 for row in review_items if row["siNo"] == item["siNo"]) > 1}
    )
    all_supplied_images = {Path(image_name).name for image_names in REAL_IMAGES.values() for image_name in image_names}
    used_source_images = {
        Path(image["sourceFile"]).name
        for row in review_items
        for image in row.get("images", [])
        if image.get("sourceFile")
    }
    unique_asset_paths = {
        image["src"]
        for row in review_items
        for image in row.get("images", [])
        if image.get("src")
    }
    missing_image_items = [
        row["name"]
        for row in review_items
        if not row.get("images")
    ]

    top_references = sorted(
        references,
        key=lambda row: ((row["downloadsNumeric"] or 0) + (row["likesNumeric"] or 0), row["name"]),
        reverse=True,
    )[:10]

    return {
        "generatedAt": datetime.now().strftime("%b %d, %Y"),
        "sourceWorkbook": str(SOURCE_XLSX),
        "deviceCount": len(devices),
        "referenceCount": len(references),
        "reviewItemCount": len(review_items),
        "categoryCounts": dict(category_counts),
        "divisionCounts": {division: division_counts.get(division, 0) for division in DIVISIONS},
        "readinessCounts": {
            "Production Ready": readiness_counts.get("Production Ready", 0),
            "Prototype Ready": readiness_counts.get("Prototype Ready", 0),
            "Design Ready": readiness_counts.get("Design Ready", 0),
            "Needs Design": readiness_counts.get("Needs Design", 0),
            "Reference Review": readiness_counts.get("Reference Review", 0),
        },
        "byCategoryDivision": by_category_division,
        "topReferences": top_references,
        "qa": {
            "sourceDashboardTotal": source_dashboard_total,
            "currentDeviceRows": len(devices),
            "currentReviewRows": len(review_items),
            "additionalImageRows": len(review_items) - len(devices),
            "duplicateRowsRemoved": duplicate_rows_removed,
            "missingTargetSettings": sum(1 for item in review_items if not item["targetSetting"]),
            "sourceDuplicateSiNumbers": source_duplicate_ids,
            "displayDuplicateSiNumbers": display_duplicate_ids,
            "missingImageItems": missing_image_items,
            "devicesWithRealImages": sum(1 for device in devices if device["images"]),
            "reviewItemsWithRealImages": with_images,
            "assignedDeviceImages": assigned_images,
            "referencesWithRealImages": reference_with_images,
            "uniqueImageAssets": len(unique_asset_paths),
            "totalImageReferences": sum(len(row.get("images", [])) for row in review_items),
            "unusedSuppliedImages": sorted(all_supplied_images - used_source_images),
        },
    }


HTML_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>STRIDE Device Review Dashboard</title>
  <link rel="icon" href="data:,">
  <style>
    :root {
      --bg: #f5f7fa;
      --surface: #ffffff;
      --surface-2: #eef3f7;
      --line: #d7e0e8;
      --line-strong: #aebdca;
      --text: #12202c;
      --muted: #617184;
      --teal: #0f766e;
      --teal-soft: #d9f3ef;
      --blue: #2563eb;
      --blue-soft: #dbeafe;
      --green: #15803d;
      --green-soft: #dcfce7;
      --amber: #b45309;
      --amber-soft: #fef3c7;
      --red: #b91c1c;
      --red-soft: #fee2e2;
      --purple: #6d28d9;
      --slate: #64748b;
      --shadow: 0 16px 34px rgba(18, 32, 44, 0.08);
      --radius: 8px;
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }

    * { box-sizing: border-box; }
    body { margin: 0; background: var(--bg); color: var(--text); line-height: 1.45; }
    button, input, select, textarea { font: inherit; }
    button { cursor: pointer; }

    .app { max-width: 1680px; margin: 0 auto; padding: 18px; }
    .topbar {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 16px;
      align-items: center;
      padding: 8px 0 14px;
    }
    .brand h1 { margin: 0; font-size: clamp(24px, 2vw, 34px); line-height: 1.08; letter-spacing: 0; }
    .brand p { margin: 6px 0 0; color: var(--muted); font-size: 14px; }
    .top-actions { display: flex; gap: 8px; flex-wrap: wrap; justify-content: flex-end; }
    .button {
      border: 1px solid var(--line);
      background: var(--surface);
      color: var(--text);
      border-radius: 7px;
      padding: 9px 12px;
      font-size: 13px;
      font-weight: 720;
      min-height: 38px;
      display: inline-flex;
      align-items: center;
      gap: 7px;
    }
    .button:hover { border-color: var(--line-strong); }
    .button.primary { background: var(--teal); color: #fff; border-color: var(--teal); }
    .button.success { background: var(--green); color: #fff; border-color: var(--green); }
    .button.warning { background: var(--amber); color: #fff; border-color: var(--amber); }
    .button.danger { background: var(--red); color: #fff; border-color: var(--red); }
    .button.ghost { background: #f8fbfd; }
    .button.small { min-height: 32px; padding: 7px 9px; font-size: 12px; }
    .icon { width: 16px; height: 16px; display: inline-block; color: currentColor; }

    .layout {
      display: grid;
      grid-template-columns: 276px minmax(0, 1fr) 390px;
      gap: 14px;
      align-items: start;
    }
    .layout > * { min-width: 0; }
    .sidebar, .review-panel { position: sticky; top: 14px; }
    .panel {
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: 0 1px 0 rgba(18, 32, 44, 0.03);
      overflow: hidden;
    }
    .section-title {
      margin: 0;
      font-size: 12px;
      color: var(--muted);
      font-weight: 820;
      letter-spacing: 0.07em;
      text-transform: uppercase;
    }
    .panel-pad { padding: 14px; }
    .panel + .panel { margin-top: 12px; }
    .metric-list { display: grid; gap: 8px; margin-top: 11px; }
    .metric {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 8px;
      padding: 10px;
      border: 1px solid var(--line);
      border-radius: 7px;
      background: #fbfdfe;
    }
    .metric-label { color: var(--muted); font-size: 12px; font-weight: 720; }
    .metric-value { font-size: 22px; font-weight: 820; line-height: 1; }
    .metric-note { grid-column: 1 / -1; color: var(--muted); font-size: 12px; }

    .filter-stack { display: grid; gap: 9px; margin-top: 12px; }
    .search, .select, .notes {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 7px;
      background: var(--surface);
      color: var(--text);
    }
    .search, .select { min-height: 38px; padding: 8px 10px; }
    .notes { min-height: 78px; padding: 9px; resize: vertical; }

    .division-button, .decision-button {
      width: 100%;
      display: grid;
      grid-template-columns: 10px minmax(0, 1fr) auto;
      align-items: center;
      gap: 8px;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      border-radius: 7px;
      padding: 8px 9px;
      font-size: 13px;
      text-align: left;
    }
    .division-button.active, .decision-button.active { border-color: var(--teal); box-shadow: 0 0 0 2px rgba(15, 118, 110, 0.12); }
    .dot { width: 10px; height: 10px; border-radius: 2px; }
    .dot.learning { background: var(--teal); }
    .dot.elderly { background: #ca8a04; }
    .dot.blind { background: var(--blue); }
    .dot.montessori { background: var(--purple); }
    .dot.others { background: var(--slate); }
    .dot.next { background: var(--green); }
    .dot.pending { background: var(--amber); }
    .dot.dispose { background: var(--red); }
    .dot.review { background: var(--slate); }

    .quality-strip {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
      margin-bottom: 14px;
    }
    .quality-item {
      background: #fffaf0;
      border: 1px solid #f3d49a;
      border-radius: var(--radius);
      color: #6b3d03;
      padding: 10px;
      font-size: 12px;
    }
    .quality-item strong { display: block; color: #4d2c00; font-size: 14px; margin-bottom: 2px; }

    .analytics-grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 14px;
      margin-bottom: 14px;
    }
    .panel-header { padding: 14px 15px 8px; }
    .panel-title { margin: 0; font-size: 16px; letter-spacing: 0; }
    .panel-subtitle { margin: 4px 0 0; color: var(--muted); font-size: 13px; }
    .panel-body { padding: 8px 15px 15px; }
    .stack-row, .funnel-row {
      display: grid;
      grid-template-columns: 132px minmax(0, 1fr) 38px;
      align-items: center;
      gap: 11px;
      margin: 12px 0;
      font-size: 13px;
    }
    .stack-label { font-weight: 720; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
    .stack-total { color: var(--muted); text-align: right; font-weight: 720; }
    .stack-track, .funnel-track {
      display: flex;
      height: 23px;
      overflow: hidden;
      border-radius: 5px;
      background: var(--surface-2);
      border: 1px solid var(--line);
    }
    .stack-seg { min-width: 2px; }
    .funnel-fill {
      height: 100%;
      display: flex;
      align-items: center;
      justify-content: flex-end;
      padding-right: 8px;
      color: #fff;
      font-size: 12px;
      font-weight: 820;
      min-width: 34px;
    }
    .legend { display: flex; flex-wrap: wrap; gap: 8px 12px; margin-top: 12px; color: var(--muted); font-size: 12px; }
    .legend span { display: inline-flex; align-items: center; gap: 6px; }

    .catalogue-panel .catalogue-toolbar {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      align-items: center;
      padding: 11px 15px;
      border-top: 1px solid var(--line);
      border-bottom: 1px solid var(--line);
      background: #fbfdfe;
    }
    .product-summary { color: var(--muted); font-size: 13px; }
    .view-toggle { display: flex; gap: 6px; }
    .product-grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 11px;
      padding: 14px 15px 15px;
    }
    .product-card {
      border: 1px solid var(--line);
      border-radius: var(--radius);
      background: #fff;
      min-width: 0;
      overflow: hidden;
      cursor: pointer;
      transition: border-color 0.15s ease, box-shadow 0.15s ease, transform 0.15s ease;
    }
    .product-card:hover, .product-card.selected {
      border-color: var(--teal);
      box-shadow: 0 10px 24px rgba(15, 118, 110, 0.12);
      transform: translateY(-1px);
    }
    .product-card img {
      display: block;
      width: 100%;
      aspect-ratio: 1.35 / 1;
      object-fit: contain;
      background: #fff;
    }
    .product-content { padding: 10px; }
    .product-name {
      min-height: 42px;
      font-size: 14px;
      font-weight: 780;
      line-height: 1.25;
      display: -webkit-box;
      -webkit-line-clamp: 2;
      -webkit-box-orient: vertical;
      overflow: hidden;
    }
    .chips { display: flex; flex-wrap: wrap; gap: 5px; margin-top: 9px; }
    .chip {
      border-radius: 999px;
      padding: 4px 8px;
      font-size: 11px;
      font-weight: 760;
      line-height: 1;
      white-space: nowrap;
      background: #edf2f7;
      color: #475569;
    }
    .chip.assistive { background: var(--teal-soft); color: #0f5f59; }
    .chip.cognitive { background: var(--blue-soft); color: #1d4ed8; }
    .chip.production, .chip.next { background: var(--green-soft); color: var(--green); }
    .chip.prototype, .chip.pending { background: var(--amber-soft); color: var(--amber); }
    .chip.design { background: #eef2ff; color: #4338ca; }
    .chip.dispose, .chip.qa { background: var(--red-soft); color: var(--red); }
    .target-line { margin-top: 8px; color: var(--muted); font-size: 12px; line-height: 1.35; min-height: 32px; }

    .list-mode .product-grid { grid-template-columns: 1fr; }
    .list-mode .product-card { display: grid; grid-template-columns: 180px minmax(0, 1fr); }
    .list-mode .product-card img { height: 132px; aspect-ratio: auto; }
    .list-mode .product-name { min-height: 0; -webkit-line-clamp: 1; }

    .selected-image {
      width: 100%;
      aspect-ratio: 1.28 / 1;
      object-fit: contain;
      display: block;
      background: #fff;
      border-bottom: 1px solid var(--line);
    }
    .selected-thumbs {
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      gap: 6px;
      padding: 8px;
      border-bottom: 1px solid var(--line);
      background: #fbfdfe;
    }
    .thumb {
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 0;
      background: #fff;
      overflow: hidden;
    }
    .thumb.active { border-color: var(--teal); box-shadow: 0 0 0 2px rgba(15, 118, 110, 0.14); }
    .thumb img { width: 100%; aspect-ratio: 1.35 / 1; object-fit: contain; display: block; background: #fff; }
    .selected-body { padding: 15px; }
    .selected-title { margin: 0; font-size: 20px; line-height: 1.18; letter-spacing: 0; }
    .detail-list { display: grid; gap: 9px; margin-top: 14px; }
    .detail-row {
      display: grid;
      grid-template-columns: 112px minmax(0, 1fr);
      gap: 10px;
      border-top: 1px solid var(--line);
      padding-top: 9px;
      font-size: 13px;
    }
    .detail-row dt { color: var(--muted); font-weight: 760; }
    .detail-row dd { margin: 0; min-width: 0; overflow-wrap: anywhere; }
    .action-grid { display: grid; grid-template-columns: 1fr; gap: 8px; margin-top: 14px; }
    .review-note { margin-top: 10px; color: var(--muted); font-size: 12px; }
    .checklist { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 13px; }
    .check-item {
      border: 1px solid var(--line);
      border-radius: 7px;
      padding: 8px;
      font-size: 12px;
      color: var(--muted);
      background: #fbfdfe;
    }
    .check-item label { display: flex; gap: 6px; align-items: flex-start; color: var(--text); font-weight: 730; }

    .tables-grid { display: grid; grid-template-columns: 1fr; gap: 14px; margin-top: 14px; }
    table { width: 100%; border-collapse: collapse; font-size: 13px; }
    th {
      text-align: left;
      color: var(--muted);
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      border-bottom: 1px solid var(--line);
      padding: 9px 8px;
      white-space: nowrap;
      overflow-wrap: anywhere;
    }
    td { border-bottom: 1px solid var(--line); padding: 10px 8px; vertical-align: top; overflow-wrap: anywhere; }
    tbody tr:hover { background: #f8fbfc; }
    .num { text-align: right; font-variant-numeric: tabular-nums; }
    .ref-thumb { width: 54px; height: 40px; object-fit: cover; border-radius: 5px; margin-right: 7px; vertical-align: middle; border: 1px solid var(--line); }
    .source-section { margin-top: 14px; padding: 15px; color: var(--muted); font-size: 13px; }
    .source-section h2 { margin: 0 0 8px; color: var(--text); font-size: 15px; }
    .source-section p { overflow-wrap: anywhere; }
    .empty { padding: 30px 16px; text-align: center; color: var(--muted); font-size: 14px; grid-column: 1 / -1; }

    @media (max-width: 1320px) {
      .layout { grid-template-columns: 250px minmax(0, 1fr); }
      .review-panel { position: static; grid-column: 1 / -1; }
      .product-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 940px) {
      .layout { grid-template-columns: 1fr; }
      .sidebar { position: static; }
      .analytics-grid, .quality-strip, .tables-grid { grid-template-columns: 1fr; }
      .product-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 640px) {
      .app { padding: 12px; }
      .topbar { grid-template-columns: 1fr; }
      .top-actions { justify-content: flex-start; }
      .product-grid { grid-template-columns: 1fr; }
      .list-mode .product-card { display: block; }
      .list-mode .product-card img { height: auto; aspect-ratio: 1.35 / 1; }
      .stack-row, .funnel-row { grid-template-columns: 108px minmax(0, 1fr) 34px; }
      .detail-row { grid-template-columns: 92px minmax(0, 1fr); }
      .checklist { grid-template-columns: 1fr; }
      table { table-layout: fixed; }
      th { white-space: normal; }
    }
  </style>
</head>
<body>
  <main class="app">
    <header class="topbar">
      <div class="brand">
        <h1>STRIDE Device Review Dashboard</h1>
        <p>Real-image product catalogue review. Source refreshed: <span id="generatedAt"></span>.</p>
      </div>
      <div class="top-actions">
        <button class="button" id="resetFilters" title="Reset dashboard filters">
          <svg class="icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 12a9 9 0 1 0 3-6.7"/><path d="M3 3v6h6"/></svg>
          Reset
        </button>
        <button class="button primary" id="downloadCsv" title="Download filtered catalogue rows">
          <svg class="icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg>
          CSV
        </button>
      </div>
    </header>

    <section class="layout">
      <aside class="sidebar">
        <section class="panel panel-pad">
          <h2 class="section-title">Catalogue</h2>
          <div class="metric-list" id="metricList"></div>
        </section>

        <section class="panel panel-pad">
          <h2 class="section-title">Filters</h2>
          <div class="filter-stack">
            <input class="search" id="searchInput" type="search" placeholder="Search product or setting">
            <select class="select" id="categoryFilter" title="Filter by category"></select>
            <select class="select" id="readinessFilter" title="Filter by readiness"></select>
            <select class="select" id="decisionFilter" title="Filter by review decision"></select>
          </div>
        </section>

        <section class="panel panel-pad">
          <h2 class="section-title">Divisions</h2>
          <div class="filter-stack" id="divisionButtons"></div>
        </section>

        <section class="panel panel-pad">
          <h2 class="section-title">Review Queue</h2>
          <div class="filter-stack" id="decisionButtons"></div>
        </section>
      </aside>

      <section>
        <section class="quality-strip" id="qualityStrip"></section>

        <section class="analytics-grid">
          <article class="panel">
            <div class="panel-header">
              <h2 class="panel-title">Category and Division Mix</h2>
              <p class="panel-subtitle">Assistive and cognitive devices across the five review divisions.</p>
            </div>
            <div class="panel-body" id="divisionChart"></div>
          </article>
          <article class="panel">
            <div class="panel-header">
              <h2 class="panel-title">Readiness Pipeline</h2>
              <p class="panel-subtitle">Highest achieved catalogue stage before reviewer decision.</p>
            </div>
            <div class="panel-body" id="readinessChart"></div>
          </article>
        </section>

        <section class="panel catalogue-panel" id="cataloguePanel">
          <div class="panel-header">
            <h2 class="panel-title">Product Review Catalogue</h2>
            <p class="panel-subtitle">Real supplied images, product context, QA flags, and review decision status.</p>
          </div>
          <div class="catalogue-toolbar">
            <div class="product-summary" id="productSummary"></div>
            <div class="view-toggle">
              <button class="button small ghost" id="gridView" title="Grid view">Grid</button>
              <button class="button small ghost" id="listView" title="List view">List</button>
            </div>
          </div>
          <div class="product-grid" id="productGrid"></div>
        </section>

        <section class="tables-grid">
          <article class="panel">
            <div class="panel-header">
              <h2 class="panel-title">Review Table</h2>
              <p class="panel-subtitle">Filtered list of Excel devices and supplied image references for quick review and export.</p>
            </div>
            <div class="panel-body">
              <table>
                <thead><tr><th>SI</th><th>Device</th><th>Division</th><th>Decision</th></tr></thead>
                <tbody id="reviewRows"></tbody>
              </table>
            </div>
          </article>
        </section>

        <section class="panel source-section">
          <h2>Source and Methodology</h2>
          <p id="sourceMethod"></p>
        </section>
      </section>

      <aside class="panel review-panel" id="selectedPanel"></aside>
    </section>
  </main>

  <script>
    const payload = __DATA_JSON__;
    const sourceDevices = payload.devices;
    const devices = payload.catalogueItems || payload.devices;
    const references = payload.references;
    const metrics = payload.metrics;
    const divisions = ["Learning", "Elderly Care", "Blind", "Montessori Kits", "Others"];
    const divisionClasses = {
      "Learning": "learning",
      "Elderly Care": "elderly",
      "Blind": "blind",
      "Montessori Kits": "montessori",
      "Others": "others"
    };
    const state = {
      search: "",
      category: "All Categories",
      division: "All Divisions",
      readiness: "All Stages",
      decision: "All Decisions",
      selectedId: devices.length ? devices[0].id : "",
      view: localStorage.getItem("strideViewMode") || "grid",
      selectedImage: 0,
      reviews: JSON.parse(localStorage.getItem("strideReviewDecisions") || "{}"),
      notes: JSON.parse(localStorage.getItem("strideReviewNotes") || "{}"),
      checklist: JSON.parse(localStorage.getItem("strideReviewChecklist") || "{}")
    };

    const els = {
      generatedAt: document.getElementById("generatedAt"),
      metricList: document.getElementById("metricList"),
      qualityStrip: document.getElementById("qualityStrip"),
      divisionChart: document.getElementById("divisionChart"),
      readinessChart: document.getElementById("readinessChart"),
      productGrid: document.getElementById("productGrid"),
      productSummary: document.getElementById("productSummary"),
      selectedPanel: document.getElementById("selectedPanel"),
      reviewRows: document.getElementById("reviewRows"),
      sourceMethod: document.getElementById("sourceMethod"),
      searchInput: document.getElementById("searchInput"),
      categoryFilter: document.getElementById("categoryFilter"),
      readinessFilter: document.getElementById("readinessFilter"),
      decisionFilter: document.getElementById("decisionFilter"),
      divisionButtons: document.getElementById("divisionButtons"),
      decisionButtons: document.getElementById("decisionButtons"),
      resetFilters: document.getElementById("resetFilters"),
      downloadCsv: document.getElementById("downloadCsv"),
      cataloguePanel: document.getElementById("cataloguePanel"),
      gridView: document.getElementById("gridView"),
      listView: document.getElementById("listView")
    };

    function escapeHTML(value) {
      return String(value ?? "").replace(/[&<>"']/g, c => ({ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#039;" })[c]);
    }
    function formatNumber(value) { return Number(value || 0).toLocaleString("en-US"); }
    function decisionFor(item) { return state.reviews[item.id] || "In Review"; }
    function decisionClass(decision) {
      return { "Next Stage": "next", "Pending": "pending", "Dispose": "dispose", "In Review": "review" }[decision] || "review";
    }
    function chipClass(value) {
      if (value === "Assistive Devices") return "assistive";
      if (value === "Cognitive Devices") return "cognitive";
      if (value === "Production Ready") return "production";
      if (value === "Prototype Ready") return "prototype";
      if (value === "Design Ready") return "design";
      return decisionClass(value);
    }
    function primaryImage(item) {
      return item.images && item.images.length ? item.images[Math.min(state.selectedImage, item.images.length - 1)].src : "";
    }
    function saveReviews() {
      localStorage.setItem("strideReviewDecisions", JSON.stringify(state.reviews));
      localStorage.setItem("strideReviewNotes", JSON.stringify(state.notes));
      localStorage.setItem("strideReviewChecklist", JSON.stringify(state.checklist));
    }
    function filteredDevices() {
      const q = state.search.trim().toLowerCase();
      return devices.filter(item => {
        const decision = decisionFor(item);
        const haystack = [item.name, item.categoryLabel, item.division, item.readiness, item.targetSetting, item.reviewFocus, item.about, decision].join(" ").toLowerCase();
        if (q && !haystack.includes(q)) return false;
        if (state.category !== "All Categories" && item.categoryLabel !== state.category) return false;
        if (state.division !== "All Divisions" && item.division !== state.division) return false;
        if (state.readiness !== "All Stages" && item.readiness !== state.readiness) return false;
        if (state.decision !== "All Decisions" && decision !== state.decision) return false;
        return true;
      });
    }
    function setSelectOptions(select, options, selected) {
      select.innerHTML = options.map(option => '<option value="' + escapeHTML(option) + '"' + (option === selected ? " selected" : "") + '>' + escapeHTML(option) + '</option>').join("");
    }
    function countsByDecision() {
      const counts = { "In Review": 0, "Next Stage": 0, "Pending": 0, "Dispose": 0 };
      devices.forEach(item => { counts[decisionFor(item)] = (counts[decisionFor(item)] || 0) + 1; });
      return counts;
    }
    function renderSidebar() {
      const dCounts = countsByDecision();
      els.metricList.innerHTML = [
        ["Review Items", metrics.reviewItemCount, "Excel devices plus supplied image references"],
        ["Assistive", metrics.categoryCounts["Assistive Devices"] || 0, "Assistive device category"],
        ["Cognitive", metrics.categoryCounts["Cognitive Devices"] || 0, "Cognitive device category"],
        ["Real Images", metrics.qa.reviewItemsWithRealImages + "/" + metrics.reviewItemCount, metrics.qa.assignedDeviceImages + " supplied image assignments"],
        ["Next Stage", dCounts["Next Stage"], "Reviewer-approved queue"]
      ].map(row => '<div class="metric"><div><div class="metric-label">' + escapeHTML(row[0]) + '</div><div class="metric-note">' + escapeHTML(row[2]) + '</div></div><div class="metric-value">' + escapeHTML(row[1]) + '</div></div>').join("");

      els.divisionButtons.innerHTML = ["All Divisions"].concat(divisions).map(division => {
        const count = division === "All Divisions" ? metrics.reviewItemCount : (metrics.divisionCounts[division] || 0);
        const dot = division === "All Divisions" ? "review" : divisionClasses[division];
        return '<button class="division-button ' + (state.division === division ? "active" : "") + '" data-division="' + escapeHTML(division) + '"><i class="dot ' + dot + '"></i><span>' + escapeHTML(division) + '</span><strong>' + count + '</strong></button>';
      }).join("");
      els.divisionButtons.querySelectorAll("button").forEach(button => {
        button.addEventListener("click", () => {
          state.division = button.dataset.division;
          renderAll();
        });
      });

      els.decisionButtons.innerHTML = ["In Review", "Next Stage", "Pending", "Dispose"].map(decision => {
        return '<button class="decision-button ' + (state.decision === decision ? "active" : "") + '" data-decision="' + escapeHTML(decision) + '"><i class="dot ' + decisionClass(decision) + '"></i><span>' + escapeHTML(decision) + '</span><strong>' + dCounts[decision] + '</strong></button>';
      }).join("");
      els.decisionButtons.querySelectorAll("button").forEach(button => {
        button.addEventListener("click", () => {
          state.decision = state.decision === button.dataset.decision ? "All Decisions" : button.dataset.decision;
          setSelectOptions(els.decisionFilter, ["All Decisions", "In Review", "Next Stage", "Pending", "Dispose"], state.decision);
          renderAll();
        });
      });
    }
    function renderQuality() {
      const qa = metrics.qa;
      const sourceDuplicateText = qa.sourceDuplicateSiNumbers.length ? "Source SI " + qa.sourceDuplicateSiNumbers.join(", ") + " normalized in dashboard." : "Source SI numbers are unique.";
      const displayDuplicateText = qa.displayDuplicateSiNumbers.length ? "Duplicate display IDs: " + qa.displayDuplicateSiNumbers.join(", ") : "No duplicate dashboard IDs.";
      const imageText = qa.missingImageItems.length ? "Missing image for " + qa.missingImageItems.length + " item(s)." : "No missing images; " + qa.uniqueImageAssets + " unique real image assets used.";
      const unusedText = qa.unusedSuppliedImages.length ? qa.unusedSuppliedImages.length + " supplied image file(s) unused." : "All supplied real image files are represented.";
      const items = [
        ["Catalogue Count", "Workbook dashboard shows " + qa.sourceDashboardTotal + "; Excel devices " + qa.currentDeviceRows + "; added image references " + qa.additionalImageRows + "; total review items " + qa.currentReviewRows + "."],
        ["Images", imageText + " " + unusedText],
        ["Target Setting", qa.missingTargetSettings + " rows need target setting review."],
        ["Catalogue ID", displayDuplicateText + " " + sourceDuplicateText]
      ];
      els.qualityStrip.innerHTML = items.map(item => '<div class="quality-item"><strong>' + escapeHTML(item[0]) + '</strong>' + escapeHTML(item[1]) + '</div>').join("");
    }
    function renderDivisionChart() {
      const categories = Object.keys(metrics.byCategoryDivision);
      const maxTotal = Math.max(...categories.map(category => divisions.reduce((sum, division) => sum + (metrics.byCategoryDivision[category][division] || 0), 0)));
      els.divisionChart.innerHTML = categories.map(category => {
        const total = divisions.reduce((sum, division) => sum + (metrics.byCategoryDivision[category][division] || 0), 0);
        const segments = divisions.map(division => {
          const count = metrics.byCategoryDivision[category][division] || 0;
          const width = total ? (count / total * 100) : 0;
          return '<div class="stack-seg" style="width:' + width + '%; background:var(--' + (division === "Elderly Care" ? "amber" : division === "Blind" ? "blue" : division === "Montessori Kits" ? "purple" : division === "Others" ? "slate" : "teal") + ')" title="' + escapeHTML(division + ": " + count) + '"></div>';
        }).join("");
        return '<div class="stack-row"><div class="stack-label">' + escapeHTML(category) + '</div><div class="stack-track" style="max-width:' + (total / maxTotal * 100) + '%">' + segments + '</div><div class="stack-total">' + total + '</div></div>';
      }).join("") + '<div class="legend">' + divisions.map(division => '<span><i class="dot ' + divisionClasses[division] + '"></i>' + escapeHTML(division) + '</span>').join("") + '</div>';
    }
    function renderReadinessChart() {
      const stages = ["Production Ready", "Prototype Ready", "Design Ready", "Needs Design", "Reference Review"];
      const colors = { "Production Ready": "#15803d", "Prototype Ready": "#b45309", "Design Ready": "#4338ca", "Needs Design": "#64748b", "Reference Review": "#0f766e" };
      const max = Math.max(...stages.map(stage => metrics.readinessCounts[stage] || 0), 1);
      els.readinessChart.innerHTML = stages.map(stage => {
        const count = metrics.readinessCounts[stage] || 0;
        const pct = count ? Math.max(7, count / max * 100) : 7;
        return '<div class="funnel-row"><div class="stack-label">' + escapeHTML(stage) + '</div><div class="funnel-track"><div class="funnel-fill" style="width:' + pct + '%; background:' + colors[stage] + '">' + count + '</div></div><div class="stack-total">' + Math.round(count / metrics.reviewItemCount * 100) + '%</div></div>';
      }).join("");
    }
    function renderProducts() {
      const rows = filteredDevices();
      if (!rows.some(item => item.id === state.selectedId) && rows.length) {
        state.selectedId = rows[0].id;
        state.selectedImage = 0;
      }
      els.cataloguePanel.classList.toggle("list-mode", state.view === "list");
      els.productSummary.textContent = formatNumber(rows.length) + " of " + formatNumber(devices.length) + " review items shown";
      if (!rows.length) {
        els.productGrid.innerHTML = '<div class="empty">No catalogue items match the current filters.</div>';
        renderSelected(null);
        renderReviewRows(rows);
        return;
      }
      els.productGrid.innerHTML = rows.map(item => {
        const decision = decisionFor(item);
        const image = item.images.length ? item.images[0].src : "";
        const qa = item.qaFlags.length ? '<span class="chip qa">QA</span>' : "";
        return '<article class="product-card ' + (item.id === state.selectedId ? "selected" : "") + '" data-id="' + escapeHTML(item.id) + '">'
          + '<img src="' + escapeHTML(image) + '" alt="' + escapeHTML(item.name) + ' real product image">'
          + '<div class="product-content">'
          + '<div class="product-name">' + escapeHTML(item.name) + '</div>'
          + '<div class="chips">'
          + '<span class="chip ' + chipClass(item.categoryLabel) + '">' + escapeHTML(item.categoryLabel.replace(" Devices", "")) + '</span>'
          + '<span class="chip">' + escapeHTML(item.division) + '</span>'
          + '<span class="chip ' + chipClass(item.readiness) + '">' + escapeHTML(item.readiness.replace(" Ready", "")) + '</span>'
          + '<span class="chip ' + decisionClass(decision) + '">' + escapeHTML(decision) + '</span>'
          + qa
          + '</div>'
          + '<div class="target-line">' + escapeHTML(item.targetSetting || "Target setting not assigned") + '</div>'
          + '</div>'
          + '</article>';
      }).join("");
      els.productGrid.querySelectorAll(".product-card").forEach(card => {
        card.addEventListener("click", () => {
          state.selectedId = card.dataset.id;
          state.selectedImage = 0;
          renderProducts();
        });
      });
      renderSelected(rows.find(item => item.id === state.selectedId) || rows[0]);
      renderReviewRows(rows);
    }
    function renderSelected(item) {
      if (!item) {
        els.selectedPanel.innerHTML = '<div class="empty">No device selected.</div>';
        return;
      }
      const decision = decisionFor(item);
      const checklist = state.checklist[item.id] || {};
      const flags = item.qaFlags.length ? item.qaFlags.map(flag => '<span class="chip qa">' + escapeHTML(flag) + '</span>').join("") : '<span class="chip next">No critical flags</span>';
      const thumbs = item.images.map((image, index) => '<button class="thumb ' + (index === state.selectedImage ? "active" : "") + '" data-index="' + index + '"><img src="' + escapeHTML(image.src) + '" alt="' + escapeHTML(item.name) + ' image ' + (index + 1) + '"></button>').join("");
      els.selectedPanel.innerHTML = '<img class="selected-image" src="' + escapeHTML(primaryImage(item)) + '" alt="' + escapeHTML(item.name) + ' selected real image">'
        + (item.images.length > 1 ? '<div class="selected-thumbs">' + thumbs + '</div>' : '')
        + '<div class="selected-body">'
        + '<h2 class="selected-title">' + escapeHTML(item.name) + '</h2>'
        + '<div class="chips"><span class="chip ' + chipClass(item.categoryLabel) + '">' + escapeHTML(item.categoryLabel) + '</span><span class="chip">' + escapeHTML(item.division) + '</span><span class="chip ' + chipClass(item.readiness) + '">' + escapeHTML(item.readiness) + '</span><span class="chip ' + decisionClass(decision) + '">' + escapeHTML(decision) + '</span></div>'
        + '<dl class="detail-list">'
        + '<div class="detail-row"><dt>Catalogue ID</dt><dd>' + escapeHTML(item.siNo) + (String(item.siNo) !== String(item.sourceSiNo) ? ' <span style="color:var(--muted)">(source SI ' + escapeHTML(item.sourceSiNo) + ')</span>' : '') + '</dd></div>'
        + '<div class="detail-row"><dt>About</dt><dd>' + escapeHTML(item.about) + '</dd></div>'
        + '<div class="detail-row"><dt>Target</dt><dd>' + escapeHTML(item.targetSetting || "Needs assignment") + '</dd></div>'
        + '<div class="detail-row"><dt>Next Stage</dt><dd>' + escapeHTML(item.nextStage) + '</dd></div>'
        + '<div class="detail-row"><dt>Review Focus</dt><dd>' + escapeHTML(item.reviewFocus) + '</dd></div>'
        + '<div class="detail-row"><dt>Images</dt><dd>' + escapeHTML(item.imageSource) + '<br><span style="color:var(--muted)">' + item.images.map(image => escapeHTML(image.fileName)).join(", ") + '</span></dd></div>'
        + '<div class="detail-row"><dt>QA Flags</dt><dd><div class="chips">' + flags + '</div></dd></div>'
        + '</dl>'
        + '<div class="action-grid">'
        + '<button class="button success" data-action="Next Stage"><svg class="icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M20 6 9 17l-5-5"/></svg> Move to Next Stage</button>'
        + '<button class="button warning" data-action="Pending"><svg class="icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2"/></svg> Keep Pending</button>'
        + '<button class="button danger" data-action="Dispose"><svg class="icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 6h18"/><path d="M8 6v14h8V6"/><path d="M10 10v6"/><path d="M14 10v6"/></svg> Dispose Idea</button>'
        + '</div>'
        + '<p class="review-note">Decision is saved in this browser and included in CSV export.</p>'
        + '<div class="checklist">'
        + checklistBox(item.id, "useCase", "Use-case fit", checklist.useCase)
        + checklistBox(item.id, "safety", "Safety check", checklist.safety)
        + checklistBox(item.id, "print", "Print feasibility", checklist.print)
        + checklistBox(item.id, "evidence", "Photo/evidence", checklist.evidence)
        + '</div>'
        + '<textarea class="notes" id="reviewNotes" placeholder="Reviewer notes">' + escapeHTML(state.notes[item.id] || "") + '</textarea>'
        + '</div>';
      els.selectedPanel.querySelectorAll(".thumb").forEach(button => {
        button.addEventListener("click", () => {
          state.selectedImage = Number(button.dataset.index);
          renderSelected(item);
        });
      });
      els.selectedPanel.querySelectorAll("[data-action]").forEach(button => {
        button.addEventListener("click", () => {
          state.reviews[item.id] = button.dataset.action;
          saveReviews();
          renderAll();
        });
      });
      els.selectedPanel.querySelectorAll("input[type='checkbox']").forEach(input => {
        input.addEventListener("change", () => {
          state.checklist[item.id] = state.checklist[item.id] || {};
          state.checklist[item.id][input.dataset.check] = input.checked;
          saveReviews();
        });
      });
      els.selectedPanel.querySelector("#reviewNotes").addEventListener("input", event => {
        state.notes[item.id] = event.target.value;
        saveReviews();
      });
    }
    function checklistBox(id, key, label, checked) {
      return '<div class="check-item"><label><input type="checkbox" data-check="' + key + '"' + (checked ? " checked" : "") + '> ' + escapeHTML(label) + '</label></div>';
    }
    function renderReviewRows(rows) {
      els.reviewRows.innerHTML = rows.slice(0, 16).map(row => '<tr><td>' + escapeHTML(row.siNo) + '</td><td><strong>' + escapeHTML(row.name) + '</strong><br><span style="color:var(--muted)">' + escapeHTML(row.targetSetting || "Target setting not assigned") + '</span></td><td>' + escapeHTML(row.division) + '</td><td><span class="chip ' + decisionClass(decisionFor(row)) + '">' + escapeHTML(decisionFor(row)) + '</span></td></tr>').join("");
    }
    function renderSource() {
      els.generatedAt.textContent = metrics.generatedAt;
      els.sourceMethod.textContent = "Source workbook: " + metrics.sourceWorkbook + ". Product Review Catalogue includes Device Catalogue rows, MakerWorld Reference rows, and supplemental supplied screenshots that were not represented in the workbook. Product images are copied from the real files supplied in Downloads into dashboard/assets/products, de-duplicated by image content, and reused where the same real source supports more than one product. Duplicate source SI numbers are normalized into unique dashboard catalogue IDs. Review decisions, notes, and checklist states are saved in browser localStorage and exported through CSV.";
    }
    function renderAll() {
      renderSidebar();
      renderQuality();
      renderDivisionChart();
      renderReadinessChart();
      renderProducts();
      renderSource();
    }
    function wireControls() {
      setSelectOptions(els.categoryFilter, ["All Categories"].concat(Array.from(new Set(devices.map(d => d.categoryLabel))).sort()), state.category);
      setSelectOptions(els.readinessFilter, ["All Stages", "Production Ready", "Prototype Ready", "Design Ready", "Needs Design", "Reference Review"], state.readiness);
      setSelectOptions(els.decisionFilter, ["All Decisions", "In Review", "Next Stage", "Pending", "Dispose"], state.decision);
      els.searchInput.addEventListener("input", event => { state.search = event.target.value; renderProducts(); });
      els.categoryFilter.addEventListener("change", event => { state.category = event.target.value; renderAll(); });
      els.readinessFilter.addEventListener("change", event => { state.readiness = event.target.value; renderAll(); });
      els.decisionFilter.addEventListener("change", event => { state.decision = event.target.value; renderAll(); });
      els.gridView.addEventListener("click", () => { state.view = "grid"; localStorage.setItem("strideViewMode", state.view); renderProducts(); });
      els.listView.addEventListener("click", () => { state.view = "list"; localStorage.setItem("strideViewMode", state.view); renderProducts(); });
      els.resetFilters.addEventListener("click", () => {
        state.search = "";
        state.category = "All Categories";
        state.division = "All Divisions";
        state.readiness = "All Stages";
        state.decision = "All Decisions";
        els.searchInput.value = "";
        wireControls();
        renderAll();
      });
      els.downloadCsv.addEventListener("click", downloadCsv);
    }
    function downloadCsv() {
      const headers = ["Catalogue ID", "Source SI No.", "Device Name", "Category", "Division", "Readiness", "Next Stage", "Review Decision", "Target Setting", "Review Focus", "Image Source", "Image Files", "Reviewer Notes", "QA Flags"];
      const rows = filteredDevices().map(row => [
        row.siNo, row.sourceSiNo, row.name, row.categoryLabel, row.division, row.readiness, row.nextStage, decisionFor(row), row.targetSetting, row.reviewFocus, row.imageSource, row.images.map(image => image.fileName).join("; "), state.notes[row.id] || "", row.qaFlags.join("; ")
      ]);
      const csv = [headers].concat(rows).map(row => row.map(cell => '"' + String(cell ?? "").replace(/"/g, '""') + '"').join(",")).join("\\n");
      const blob = new Blob([csv], { type: "text/csv;charset=utf-8" });
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      link.download = "stride_product_review_decisions.csv";
      document.body.appendChild(link);
      link.click();
      link.remove();
      URL.revokeObjectURL(url);
    }
    function init() {
      wireControls();
      renderAll();
    }
    init();
  </script>
</body>
</html>
"""


def build() -> None:
    COPIED_IMAGE_BY_HASH.clear()
    if OLD_GENERATED_CONCEPT.exists():
        OLD_GENERATED_CONCEPT.unlink()
    if SOURCE_XLSX.exists() and PRODUCT_ASSET_DIR.exists():
        shutil.rmtree(PRODUCT_ASSET_DIR)
    devices, references, review_items, metrics = load_catalogue()
    OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    payload = {"devices": devices, "references": references, "catalogueItems": review_items, "metrics": metrics}
    html = HTML_TEMPLATE.replace("__DATA_JSON__", json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(OUTPUT_HTML)
    print(f"devices={len(devices)} references={len(references)} real_images={metrics['qa']['devicesWithRealImages']}")


if __name__ == "__main__":
    build()
